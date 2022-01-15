class other_functions:
	def get_openpyxl():
		try:
			from openpyxl import load_workbook
		except:
			from os import system
			system("pip install openpyxl")
			from openpyxl import load_workbook
		return load_workbook

	def base64_file_encode(file_path: str):
		from base64 import b64encode
		with open(file_path, "rb") as file:
			bytes_str = b64encode(file.read())
		return bytes_str

	def base64_file_decode(base64_bytes: str):
		from base64 import b64decode
		decrypted = b64decode(base64_bytes)
		return decrypted

	def base64_write(decoded_bytes: str):
		import datetime
		now = datetime.datetime.now()
		filename = ("%2i%2i%2i%4i" % (now.hour, now.minute, now.second, now.year)).replace(' ', '0')
		path = './outputs/{0}.xlsx'.format(filename)
		with open(path, 'wb') as excel_file:
			excel_file.write(decoded_bytes)
		return path.replace('..', '.')

	def postion_table(sheet):
		final_status = False
		for value in range(2, 27):
			for _row, _col in zip(range(2, value), range(2, value)):
				deletion_value = value-1
				cell_name_one_check = sheet.cell(row = 1, column = 1).value
				cell_name_zer_check = sheet.cell(row = _row, column = _col).value
				cell_name_row_check = sheet.cell(row = _row, column = value).value
				cell_name_col_check = sheet.cell(row = value, column = _col).value
				if not isinstance(cell_name_one_check, type(None)):
					if cell_name_one_check.upper().replace(' ', '').startswith('TIMETABLE'):
						final_status = True
						break
				elif not isinstance(cell_name_row_check, type(None)):
					if cell_name_row_check.upper().replace(' ', '').startswith('TIMETABLE'):
						sheet.delete_rows(1, _row-1)
						sheet.delete_cols(1, deletion_value)
						final_status = True
						break
				elif not isinstance(cell_name_col_check, type(None)):
					if cell_name_col_check.upper().replace(' ', '').startswith('TIMETABLE'):
						sheet.delete_rows(1, deletion_value)
						sheet.delete_cols(1, _col-1)
						final_status = True
						break
				elif not isinstance(cell_name_zer_check, type(None)):
					if cell_name_zer_check.upper().replace(' ', '').startswith('TIMETABLE'):
						sheet.delete_rows(1, _row-1)
						sheet.delete_cols(1, _col-1)
						final_status = True
						break
				else:
					final_status = False
		return final_status

	def get_class(title: str):
		class_name0 = title.upper().replace(' ', '').replace('TIMETABLEFORCLASS', '')
		try: 
			int(class_name0[:2])
			class_name = class_name0[:3]
		except:
			class_name = class_name0[:2]
		return class_name
	
	def erase_file(file_path: str):
		import os
		try:
			os.remove(file_path)
			return True
		except:
			return False
